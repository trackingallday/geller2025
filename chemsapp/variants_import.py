"""Importer for the product-variants spreadsheet (e.g. "Variantsv2.xlsx").

Shared by the `import_variants` management command and the ProductVariant
admin upload page. Sheet layout: header on row 2, data from row 3 —
A "Website ID" (Product pk), B Code, C Product name, D Size name,
E Pack Size, F Description, G Barcode No.

Existing variants (matched by code, case-insensitive) are skipped, never
updated. Missing Sizes are created from the sheet's size names, parsing a
volume out of names like "5L" / "500ml" / "10kg Bag" so cost-in-use maths
works (kg treated 1:1 as litres, consistent with the dilutions importer).
"""
import re
from decimal import Decimal

import openpyxl

from chemsapp.models import Product, ProductVariant, Size

HEADER_ROW = 2
FIRST_DATA_ROW = 3
COL_PRODUCT_ID = 1  # A "Website ID"
COL_CODE = 2        # B
COL_PRODUCT_NAME = 3  # C
COL_SIZE = 4        # D
COL_PACK_SIZE = 5   # E
COL_DESCRIPTION = 6  # F
COL_BARCODE = 7     # G

_VOLUME_RE = re.compile(r'^(\d+(?:\.\d+)?)\s*(l|ltr|litre|ml|kg|gm|g)\b', re.IGNORECASE)
_ML_UNITS = {'ml', 'gm', 'g'}


def parse_volume_litres(size_name):
    """Best-effort container volume from a size name: '5L' → 5, '500ml' → 0.5,
    '10kg Bag' → 10 (kg≈litres fudge). None when unparseable (Aerosol, 200S…)."""
    match = _VOLUME_RE.match(size_name.strip())
    if not match:
        return None
    volume = Decimal(match.group(1))
    if match.group(2).lower() in _ML_UNITS:
        volume /= 1000
    return volume


def import_variants_workbook(file_or_path):
    """Run the import against an xlsx file (path or file-like object).

    Must be called inside a transaction (the callers own dry-run semantics).
    Returns a stats dict: created, skipped_existing, unknown_products,
    duplicate_codes, created_sizes (lists).
    """
    workbook = openpyxl.load_workbook(file_or_path, data_only=True)
    sheet = workbook.worksheets[0]

    products = Product.objects.in_bulk()
    existing_codes = {
        c.strip().upper()
        for c in ProductVariant.objects.exclude(code__isnull=True).values_list('code', flat=True)
        if c.strip()
    }
    sizes_by_name = {}
    for size in Size.objects.all():
        sizes_by_name.setdefault(size.name.strip().lower(), size)

    stats = {'created': [], 'skipped_existing': [], 'unknown_products': [],
             'duplicate_codes': [], 'created_sizes': []}
    seen_codes = set()

    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        product_id = sheet.cell(row=row, column=COL_PRODUCT_ID).value
        code_cell = sheet.cell(row=row, column=COL_CODE).value
        if product_id is None and code_cell is None:
            continue
        code = str(code_cell or '').strip()
        sheet_name = str(sheet.cell(row=row, column=COL_PRODUCT_NAME).value or '').strip()
        label = f'{code} ({sheet_name})'

        if not code:
            stats['unknown_products'].append(f'row {row} ({sheet_name}) — no code')
            continue
        if code.upper() in seen_codes:
            stats['duplicate_codes'].append(f'{label} — row {row} skipped, code already used in sheet')
            continue
        seen_codes.add(code.upper())

        if code.upper() in existing_codes:
            stats['skipped_existing'].append(label)
            continue

        product = products.get(product_id)
        if product is None:
            stats['unknown_products'].append(f'{label} — Website ID {product_id} not found')
            continue

        size = None
        size_name = str(sheet.cell(row=row, column=COL_SIZE).value or '').strip()
        if size_name:
            size = sizes_by_name.get(size_name.lower())
            if size is None:
                size = Size.objects.create(
                    name=size_name, desc=size_name, amount='',
                    volume_litres=parse_volume_litres(size_name),
                )
                sizes_by_name[size_name.lower()] = size
                volume = f'{size.volume_litres}L' if size.volume_litres is not None else 'no volume parsed'
                stats['created_sizes'].append(f'{size_name} ({volume})')

        barcode_cell = sheet.cell(row=row, column=COL_BARCODE).value
        if isinstance(barcode_cell, float) and barcode_cell.is_integer():
            barcode = str(int(barcode_cell))
        elif isinstance(barcode_cell, int):
            barcode = str(barcode_cell)
        elif isinstance(barcode_cell, str) and barcode_cell.strip().isdigit():
            barcode = barcode_cell.strip()
        else:
            barcode = ''  # '#N/A' and other placeholders

        pack_cell = sheet.cell(row=row, column=COL_PACK_SIZE).value
        try:
            pack_size = int(pack_cell)
        except (TypeError, ValueError):
            pack_size = 1

        ProductVariant.objects.create(
            code=code,
            product=product,
            size=size,
            pack_size=pack_size,
            description=str(sheet.cell(row=row, column=COL_DESCRIPTION).value or '').strip(),
            barcode=barcode,
        )
        existing_codes.add(code.upper())
        stats['created'].append(f'{label} — {product.name}, {size_name or "no size"}')

    return stats
