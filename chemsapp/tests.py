import tempfile
from decimal import Decimal
from io import StringIO

import openpyxl
from django.core.management import call_command
from django.test import TestCase

from chemsapp.models import ApplicationType, DilutionVariant, Product, ProductVariant, Size

# Column layout of the "Dilutions for Geller" sheet (see import_dilutions command)
COL_NAME, COL_SKU, COL_PACK, COL_VOL = 3, 6, 7, 41
COL_MOP_GENERAL, COL_SPRAY_HD, COL_SANITISING_200, COL_AUTODOSE = 14, 17, 28, 34


def write_sheet(rows):
    """Build a minimal Dilutions xlsx; `rows` is a list of {column: value} dicts
    written from row 5 down. Returns the file path."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, row in enumerate(rows):
        for column, value in row.items():
            sheet.cell(row=5 + i, column=column, value=value)
    path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    workbook.save(path)
    return path


def run_import(path, *flags):
    out = StringIO()
    call_command('import_dilutions', path, *flags, stdout=out)
    return out.getvalue()


def make_product(name='Konquer', code='KONQ'):
    return Product.objects.create(
        name=name, productCode=code, brand='Geller',
        description='d', directions='d',
    )


class ImportDilutionsTestCase(TestCase):
    serialized_rollback = True

    def setUp(self):
        self.size_5l = Size.objects.create(name='5L', desc='5 Litre', amount='5')
        self.product = make_product()
        self.variant = ProductVariant.objects.create(
            code='KONQ05', product=self.product, size=self.size_5l, pack_size=1, barcode='')

    def row(self, overrides=None):
        base = {
            COL_NAME: 'Konquer', COL_SKU: 'KONQ05', COL_PACK: '5L', COL_VOL: 5,
            COL_MOP_GENERAL: 100, COL_SPRAY_HD: 12.5,
        }
        base.update(overrides or {})
        return base

    def test_seeds_application_types_and_imports_values(self):
        output = run_import(write_sheet([self.row({COL_SANITISING_200: 120, COL_AUTODOSE: 'Check rec'})]))
        self.assertEqual(ApplicationType.objects.count(), 30)
        self.assertIn('Matched variants: 1', output)

        dilutions = {d.application_type.name: d for d in self.variant.dilutions.all()}
        self.assertEqual(dilutions['Floor Mop/Bucket - General Cleaning'].value, Decimal('100'))
        self.assertEqual(dilutions['Spray Bottle - Heavy Duty Cleaning'].value, Decimal('12.5'))
        sanitising = dilutions['Sanitising 200-400ppm']
        self.assertEqual(sanitising.value, Decimal('120'))
        self.assertEqual(sanitising.application_type.value_kind, ApplicationType.ML_PER_LITRE)
        autodose = dilutions['AutoDose - Per Cycle']
        self.assertIsNone(autodose.value)
        self.assertEqual(autodose.note, 'Check rec')

    def test_reimport_updates_and_deletes_stale(self):
        run_import(write_sheet([self.row()]))
        self.assertEqual(self.variant.dilutions.count(), 2)
        # Mop value changes, spray column now empty → its row must go
        run_import(write_sheet([self.row({COL_MOP_GENERAL: 50, COL_SPRAY_HD: None})]))
        dilutions = list(self.variant.dilutions.all())
        self.assertEqual(len(dilutions), 1)
        self.assertEqual(dilutions[0].value, Decimal('50'))

    def test_volume_backfilled_onto_size(self):
        self.assertIsNone(self.size_5l.volume_litres)
        run_import(write_sheet([self.row()]))
        self.size_5l.refresh_from_db()
        self.assertEqual(self.size_5l.volume_litres, Decimal('5'))

    def test_duplicate_sku_second_row_skipped(self):
        output = run_import(write_sheet([
            self.row(),
            self.row({COL_PACK: '20L', COL_MOP_GENERAL: 999}),
        ]))
        self.assertIn('Duplicate SKUs', output)
        self.assertEqual(self.variant.dilutions.get(
            application_type__name='Floor Mop/Bucket - General Cleaning').value, Decimal('100'))

    def test_unmatched_sku_reported(self):
        output = run_import(write_sheet([self.row({COL_SKU: 'NOPE05', COL_NAME: 'Nope'})]))
        self.assertIn('Unmatched SKUs (1)', output)
        self.assertIn('NOPE05', output)
        self.assertEqual(DilutionVariant.objects.count(), 0)

    def test_product_codes_freetext_fallback(self):
        # SKU appears only in the free-text productCodes field
        self.product.productCodes = 'KONQOLD, KONQ05B / KONQNEW'
        self.product.save()
        output = run_import(write_sheet([self.row({COL_SKU: 'KONQNEW'})]))
        self.assertIn('Matched variants: 1', output)
        self.assertEqual(self.variant.dilutions.count(), 2)

    def test_product_code_fallback_to_single_variant(self):
        # SKU matches Product.productCode, not the variant code; product has one variant
        output = run_import(write_sheet([self.row({COL_SKU: 'KONQ'})]))
        self.assertIn('Matched variants: 1', output)
        self.assertEqual(self.variant.dilutions.count(), 2)

    def test_fallback_variant_not_claimed_twice(self):
        run_import(write_sheet([
            self.row(),  # claims KONQ05 by code
            self.row({COL_SKU: 'KONQ', COL_MOP_GENERAL: 999}),  # falls back to same variant
        ]))
        self.assertEqual(self.variant.dilutions.get(
            application_type__name='Floor Mop/Bucket - General Cleaning').value, Decimal('100'))

    def test_create_missing_variants(self):
        # A product matched by name whose variants can't be resolved (it has none)
        other = make_product(name='Gleam', code='GLEAM')
        rows = [{COL_NAME: 'Gleam', COL_SKU: 'GLEAM20', COL_PACK: '20L', COL_MOP_GENERAL: 30, COL_VOL: 20}]

        output = run_import(write_sheet(rows))
        self.assertIn('Product matched but no variant (1)', output)
        self.assertIn('GLEAM20', output)

        output = run_import(write_sheet(rows), '--create-missing-variants')
        self.assertIn('Created variants (1)', output)
        self.assertIn('no Size matched', output)
        created = ProductVariant.objects.get(code='GLEAM20')
        self.assertEqual(created.product, other)
        self.assertIsNone(created.size)  # no '20L' Size exists
        self.assertEqual(created.dilutions.count(), 1)

    def test_create_flag_makes_new_variant_for_other_pack_size(self):
        # Product has one coded variant (KONQ05); a row for a different pack
        # (matched via product name) must create KONQ20, not hijack KONQ05.
        rows = [self.row({COL_SKU: 'KONQ20', COL_PACK: '20L', COL_MOP_GENERAL: 55})]
        run_import(write_sheet(rows), '--create-missing-variants')
        created = ProductVariant.objects.get(code='KONQ20')
        self.assertEqual(created.product, self.product)
        self.assertEqual(created.dilutions.get(
            application_type__name='Floor Mop/Bucket - General Cleaning').value, Decimal('55'))
        self.assertEqual(self.variant.dilutions.count(), 0)  # KONQ05 untouched

    def test_row_without_dilution_values_skipped(self):
        output = run_import(write_sheet([self.row({COL_MOP_GENERAL: None, COL_SPRAY_HD: None})]))
        self.assertIn('no dilution data (skipped): 1', output)
        self.assertEqual(DilutionVariant.objects.count(), 0)

    def test_dry_run_rolls_back(self):
        output = run_import(write_sheet([self.row()]), '--dry-run')
        self.assertIn('Dry run', output)
        self.assertEqual(DilutionVariant.objects.count(), 0)
        self.assertEqual(ApplicationType.objects.count(), 0)


def write_variants_sheet(rows):
    """Build a minimal variants xlsx; `rows` is a list of 7-tuples
    (website_id, code, product, size, pack_size, description, barcode)
    written from row 3 down. Returns the file path."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, values in enumerate(rows):
        for col, value in enumerate(values, start=1):
            sheet.cell(row=3 + i, column=col, value=value)
    path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    workbook.save(path)
    return path


def run_variants_import(path, *flags):
    out = StringIO()
    call_command('import_variants', path, *flags, stdout=out)
    return out.getvalue()


class ImportVariantsTestCase(TestCase):
    serialized_rollback = True

    def setUp(self):
        self.product = make_product()

    def test_creates_variant_with_new_size_and_parsed_volume(self):
        from decimal import Decimal as D
        output = run_variants_import(write_variants_sheet([
            (self.product.pk, 'KONQ05', 'Konquer', '5L', 3, 'Geller Konquer 5L', 9421033270001),
        ]))
        self.assertIn('Variants created: 1', output)
        variant = ProductVariant.objects.get(code='KONQ05')
        self.assertEqual(variant.product, self.product)
        self.assertEqual(variant.pack_size, 3)
        self.assertEqual(variant.barcode, '9421033270001')
        self.assertEqual(variant.size.name, '5L')
        self.assertEqual(variant.size.volume_litres, D('5'))

    def test_volume_parsing_variants(self):
        from chemsapp.variants_import import parse_volume_litres
        from decimal import Decimal as D
        self.assertEqual(parse_volume_litres('500ml'), D('0.5'))
        self.assertEqual(parse_volume_litres('500ml Pump'), D('0.5'))
        self.assertEqual(parse_volume_litres('10kg Bag'), D('10'))
        self.assertEqual(parse_volume_litres('3.3L'), D('3.3'))
        self.assertIsNone(parse_volume_litres('Aerosol'))
        self.assertIsNone(parse_volume_litres('200S'))

    def test_existing_code_skipped_not_updated(self):
        existing = ProductVariant.objects.create(
            code='KONQ05', product=self.product, pack_size=1, barcode='original')
        output = run_variants_import(write_variants_sheet([
            (self.product.pk, 'konq05', 'Konquer', '5L', 3, 'desc', 123),
        ]))
        self.assertIn('Skipped (already exist): 1', output)
        existing.refresh_from_db()
        self.assertEqual(existing.barcode, 'original')  # untouched
        self.assertEqual(ProductVariant.objects.count(), 1)

    def test_existing_size_reused_case_insensitively(self):
        size = Size.objects.create(name='5l', desc='5 Litre', amount='5')
        run_variants_import(write_variants_sheet([
            (self.product.pk, 'KONQ05', 'Konquer', '5L', 3, 'desc', 123),
        ]))
        self.assertEqual(ProductVariant.objects.get(code='KONQ05').size, size)
        self.assertEqual(Size.objects.count(), 1)

    def test_unknown_product_reported(self):
        output = run_variants_import(write_variants_sheet([
            (99999, 'NOPE05', 'Nope', '5L', 3, 'desc', 123),
        ]))
        self.assertIn('Unknown products (1)', output)
        self.assertIn('Website ID 99999', output)
        self.assertEqual(ProductVariant.objects.count(), 0)

    def test_duplicate_code_in_sheet_skipped(self):
        output = run_variants_import(write_variants_sheet([
            (self.product.pk, 'GLEAM05', 'Gleam', '5L', 3, 'desc', 123),
            (self.product.pk, 'GLEAM05', 'Gleam', '20L', 1, 'desc', 456),
        ]))
        self.assertIn('Duplicate codes in sheet (1)', output)
        self.assertEqual(ProductVariant.objects.count(), 1)

    def test_na_barcode_becomes_empty(self):
        run_variants_import(write_variants_sheet([
            (self.product.pk, 'KONQ05', 'Konquer', '5L', 3, 'desc', '#N/A'),
        ]))
        self.assertEqual(ProductVariant.objects.get(code='KONQ05').barcode, '')

    def test_dry_run_rolls_back(self):
        output = run_variants_import(write_variants_sheet([
            (self.product.pk, 'KONQ05', 'Konquer', '5L', 3, 'desc', 123),
        ]), '--dry-run')
        self.assertIn('Dry run', output)
        self.assertEqual(ProductVariant.objects.count(), 0)
        self.assertEqual(Size.objects.count(), 0)


class AdminImportVariantsTestCase(TestCase):
    serialized_rollback = True
    URL = '/admin/chemsapp/productvariant/import-variants/'

    def setUp(self):
        from django.contrib.auth.models import User
        self.admin = User.objects.create_superuser('boss2', 'boss2@example.com', 'pass12345')
        self.client.force_login(self.admin)
        self.product = make_product()
        self.sheet_path = write_variants_sheet([
            (self.product.pk, 'KONQ05', 'Konquer', '5L', 3, 'Geller Konquer 5L', 9421033270001),
        ])

    def test_changelist_links_to_import(self):
        response = self.client.get('/admin/chemsapp/productvariant/')
        self.assertContains(response, 'import-variants/')

    def test_import_applies_and_reports(self):
        with open(self.sheet_path, 'rb') as f:
            response = self.client.post(self.URL, {'xlsx_file': f})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Variants created:</strong> 1')
        self.assertTrue(ProductVariant.objects.filter(code='KONQ05').exists())

    def test_dry_run_reports_without_saving(self):
        with open(self.sheet_path, 'rb') as f:
            response = self.client.post(self.URL, {'xlsx_file': f, 'dry_run': 'on'})
        self.assertContains(response, 'Variants created:</strong> 1')
        self.assertFalse(ProductVariant.objects.exists())

    def test_requires_staff(self):
        self.client.logout()
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 302)


class AdminImportDilutionsTestCase(TestCase):
    serialized_rollback = True
    URL = '/admin/chemsapp/applicationtype/import-dilutions/'

    def setUp(self):
        from django.contrib.auth.models import User
        self.admin = User.objects.create_superuser('boss', 'boss@example.com', 'pass12345')
        self.client.force_login(self.admin)
        self.size_5l = Size.objects.create(name='5L', desc='5 Litre', amount='5')
        self.product = make_product()
        self.variant = ProductVariant.objects.create(
            code='KONQ05', product=self.product, size=self.size_5l, pack_size=1, barcode='')
        self.sheet_path = write_sheet([{
            COL_NAME: 'Konquer', COL_SKU: 'KONQ05', COL_PACK: '5L', COL_VOL: 5,
            COL_MOP_GENERAL: 100, COL_SPRAY_HD: 12.5,
        }])

    def _post(self, **extra):
        with open(self.sheet_path, 'rb') as f:
            return self.client.post(self.URL, {'xlsx_file': f, **extra})

    def test_requires_staff(self):
        self.client.logout()
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 302)  # redirected to admin login

    def test_page_renders_form(self):
        response = self.client.get(self.URL)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Import dilutions spreadsheet')
        self.assertContains(response, 'xlsx_file')

    def test_changelist_links_to_import(self):
        response = self.client.get('/admin/chemsapp/applicationtype/')
        self.assertContains(response, 'import-dilutions/')

    def test_import_applies_and_reports(self):
        response = self._post()
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Matched variants:</strong> 1')
        self.assertEqual(ApplicationType.objects.count(), 30)
        self.assertEqual(self.variant.dilutions.count(), 2)

    def test_dry_run_reports_without_saving(self):
        response = self._post(dry_run='on')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Matched variants:</strong> 1')
        self.assertEqual(ApplicationType.objects.count(), 0)
        self.assertEqual(DilutionVariant.objects.count(), 0)

    def test_missing_file_shows_error(self):
        response = self.client.post(self.URL, {})
        self.assertContains(response, 'Choose an .xlsx file')

    def test_invalid_file_shows_error(self):
        import tempfile
        bad = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        bad.write(b'not a spreadsheet')
        bad.close()
        with open(bad.name, 'rb') as f:
            response = self.client.post(self.URL, {'xlsx_file': f})
        self.assertContains(response, 'Import failed')
        self.assertEqual(ApplicationType.objects.count(), 0)
