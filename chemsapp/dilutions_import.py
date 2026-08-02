"""Importer for the "Dilutions for Geller" spreadsheet.

Shared by the `import_dilutions` management command and the ApplicationType
admin upload page. Seeds ApplicationTypes from the sheet's column layout,
matches each row's SKU to a ProductVariant and syncs its DilutionVariant rows.
"""
import re
from decimal import Decimal, InvalidOperation

import openpyxl

from chemsapp.models import ApplicationType, DilutionVariant, Product, ProductVariant, Size

# Layout of the "Dilutions for Geller" spreadsheet.
HEADER_ROW = 3
FIRST_DATA_ROW = 5
COL_PRODUCT_NAME = 3   # C
COL_SKU = 6            # F
COL_PACK_SIZE = 7      # G
COL_VOLUME_LITRES = 41  # AO "Vol (L)"

# (column, name, category, value_kind, unit_volume_litres, unit_label)
# One entry per application-type column J..AM; the cell value semantics
# (ratio vs ml/g dose) come from the column, not the cell.
APPLICATION_COLUMNS = [
    (10, 'Wash up Sink - Manual Dishwashing', 'Wash up Sink', ApplicationType.RATIO, '1', 'per litre'),
    (11, 'Autoscrubber - Light Duty Cleaning', 'Autoscrubber', ApplicationType.RATIO, '1', 'per litre'),
    (12, 'Floor Mop/Bucket - Stripping', 'Floor Mop/Bucket', ApplicationType.RATIO, '1', 'per litre'),
    (13, 'Floor Mop/Bucket - Heavy Duty Cleaning', 'Floor Mop/Bucket', ApplicationType.RATIO, '1', 'per litre'),
    (14, 'Floor Mop/Bucket - General Cleaning', 'Floor Mop/Bucket', ApplicationType.RATIO, '1', 'per litre'),
    (15, 'Floor Mop/Bucket - Light Duty Cleaning', 'Floor Mop/Bucket', ApplicationType.RATIO, '1', 'per litre'),
    (16, 'Applicator Bottle - Applied Directly', 'Applicator Bottle', ApplicationType.RATIO, '0.75', 'per 750ml bottle'),
    (17, 'Spray Bottle - Heavy Duty Cleaning', 'Spray Bottle', ApplicationType.RATIO, '0.75', 'per 750ml spray bottle'),
    (18, 'Spray Bottle - General Cleaning', 'Spray Bottle', ApplicationType.RATIO, '0.75', 'per 750ml spray bottle'),
    (19, 'Spray Bottle - Light Duty Cleaning', 'Spray Bottle', ApplicationType.RATIO, '0.75', 'per 750ml spray bottle'),
    (20, 'Spray Bottle - Sanitising 200-400ppm', 'Spray Bottle', ApplicationType.RATIO, '0.75', 'per 750ml spray bottle'),
    (21, 'Manual Cleaning - Heavy Duty Cleaning', 'Manual Cleaning', ApplicationType.RATIO, '1', 'per litre'),
    (22, 'Manual Cleaning - General Cleaning', 'Manual Cleaning', ApplicationType.RATIO, '1', 'per litre'),
    (23, 'Manual Cleaning - Light Duty Cleaning', 'Manual Cleaning', ApplicationType.RATIO, '1', 'per litre'),
    (24, 'Foaming Equipment - Light Duty Cleaning', 'Foaming Equipment', ApplicationType.RATIO, '1', 'per litre'),
    (25, 'Foaming Equipment - General Cleaning', 'Foaming Equipment', ApplicationType.RATIO, '1', 'per litre'),
    (26, 'Foaming Equipment - Heavy Duty Cleaning', 'Foaming Equipment', ApplicationType.RATIO, '1', 'per litre'),
    (27, 'Sanitising 50-100ppm', 'Sanitising', ApplicationType.ML_PER_LITRE, '1', 'per litre @ 50-100ppm'),
    (28, 'Sanitising 200-400ppm', 'Sanitising', ApplicationType.ML_PER_LITRE, '1', 'per litre @ 200-400ppm'),
    (29, 'Sanitising 500-1000ppm', 'Sanitising', ApplicationType.ML_PER_LITRE, '1', 'per litre @ 500-1000ppm'),
    (30, 'Sanitising 1000ppm', 'Sanitising', ApplicationType.ML_PER_LITRE, '1', 'per litre @ 1000ppm'),
    (31, 'Sanitising 2000ppm', 'Sanitising', ApplicationType.ML_PER_LITRE, '1', 'per litre @ 2000ppm'),
    (32, 'Laundry - mls per kg', 'Laundry', ApplicationType.ML_PER_KG, None, 'per kg of laundry'),
    (33, 'Laundry - grams per kg', 'Laundry', ApplicationType.G_PER_KG, None, 'per kg of laundry'),
    (34, 'AutoDose - Per Cycle', 'AutoDose', ApplicationType.ML_PER_CYCLE, None, 'per cycle'),
    (35, 'Soaking - grams per litre', 'Soaking', ApplicationType.G_PER_LITRE, '1', 'per litre'),
    (36, 'Warewashing - mls per litre', 'Warewashing', ApplicationType.ML_PER_LITRE, '1', 'per litre'),
    (37, 'Warewashing - mls per cycle', 'Warewashing', ApplicationType.ML_PER_CYCLE, None, 'per cycle'),
    (38, 'Warewashing - grams per cycle', 'Warewashing', ApplicationType.G_PER_CYCLE, None, 'per cycle'),
    (39, 'Warewashing - grams per litre', 'Warewashing', ApplicationType.G_PER_LITRE, '1', 'per litre'),
]


def import_dilutions_workbook(file_or_path, create_missing_variants=False):
    """Run the full import against an xlsx file (path or file-like object).

    Must be called inside a transaction (the callers own dry-run semantics).
    Returns a stats dict: matched, created_variants, unmatched, skipped_empty,
    duplicate_skus (lists), dilutions, volume_backfills (counts).
    """
    workbook = openpyxl.load_workbook(file_or_path, data_only=True)
    sheet = workbook.worksheets[0]
    app_types = _seed_application_types()
    return _import_rows(sheet, app_types, create_missing_variants)


def _seed_application_types():
    """Create/refresh one ApplicationType per sheet column. Returns {column: ApplicationType}."""
    app_types = {}
    for i, (column, name, category, kind, unit_volume, unit_label) in enumerate(APPLICATION_COLUMNS):
        app_type, _ = ApplicationType.objects.update_or_create(
            name=name,
            defaults={
                'category': category,
                'value_kind': kind,
                'unit_volume_litres': Decimal(unit_volume) if unit_volume else None,
                'unit_label': unit_label,
                'sort_order': i,
            },
        )
        app_types[column] = app_type
    return app_types


def _import_rows(sheet, app_types, create_missing):
    variants_by_code = {
        v.code.strip().upper(): v
        for v in ProductVariant.objects.exclude(code__isnull=True).exclude(code='').select_related('product', 'size')
    }
    products_by_code = {
        p.productCode.strip().upper(): p
        for p in Product.objects.exclude(productCode__isnull=True).exclude(productCode='')
    }
    # Product.productCodes is a free-text field that may list several codes;
    # tokens claimed by more than one product are ambiguous and dropped.
    AMBIGUOUS = object()
    for p in Product.objects.exclude(productCodes__isnull=True).exclude(productCodes=''):
        for token in re.split(r'[,;/\s]+', p.productCodes):
            token = token.strip().upper()
            if not token:
                continue
            existing = products_by_code.get(token)
            if existing is None:
                products_by_code[token] = p
            elif existing is not p:
                products_by_code[token] = AMBIGUOUS
    products_by_code = {k: v for k, v in products_by_code.items() if v is not AMBIGUOUS}
    products_by_name = {p.name.strip().lower(): p for p in Product.objects.all()}
    sizes_by_name = {s.name.strip().lower(): s for s in Size.objects.all()}

    stats = {'matched': [], 'created_variants': [], 'unmatched': [], 'skipped_empty': [],
             'dilutions': 0, 'volume_backfills': 0, 'duplicate_skus': []}
    seen_skus = set()
    claimed_variant_ids = set()

    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        product_name = sheet.cell(row=row, column=COL_PRODUCT_NAME).value
        sku_cell = sheet.cell(row=row, column=COL_SKU).value
        if not product_name or not sku_cell:
            continue
        sku = str(sku_cell).strip().upper()
        product_name = str(product_name).strip()
        pack_size = str(sheet.cell(row=row, column=COL_PACK_SIZE).value or '').strip()

        if sku in seen_skus:
            stats['duplicate_skus'].append(f'{sku} (row {row}, {product_name} {pack_size}) — row skipped')
            continue
        seen_skus.add(sku)

        cells = {col: sheet.cell(row=row, column=col).value for col, *_ in APPLICATION_COLUMNS}
        if all(v is None for v in cells.values()):
            stats['skipped_empty'].append(f'{sku} ({product_name} {pack_size})')
            continue

        variant = _resolve_variant(
            sku, product_name, pack_size, variants_by_code, products_by_code,
            products_by_name, sizes_by_name, create_missing, stats,
        )
        # A fallback-matched variant must not be claimed by two sheet rows —
        # the second row would silently overwrite the first row's dilutions.
        if variant is not None and variant.pk in claimed_variant_ids:
            variant = None
        if variant is None:
            stats['unmatched'].append(f'{sku} ({product_name} {pack_size})')
            continue
        claimed_variant_ids.add(variant.pk)
        stats['matched'].append(sku)

        _backfill_volume(variant, sheet.cell(row=row, column=COL_VOLUME_LITRES).value, stats)
        stats['dilutions'] += _sync_dilutions(variant, cells, app_types)

    return stats


def _resolve_variant(sku, product_name, pack_size, variants_by_code, products_by_code,
                     products_by_name, sizes_by_name, create_missing, stats):
    variant = variants_by_code.get(sku)
    if variant:
        return variant

    product = products_by_code.get(sku) or products_by_name.get(product_name.lower())
    if product is None:
        return None

    existing = list(product.variants.all()[:2])
    if len(existing) == 1:
        return existing[0]
    if not create_missing:
        return None

    variant = ProductVariant.objects.create(
        code=sku,
        product=product,
        size=sizes_by_name.get(pack_size.lower()),
        pack_size=1,
        barcode='',
    )
    variants_by_code[sku] = variant
    stats['created_variants'].append(f'{sku} ({product.name} {pack_size})')
    return variant


def _backfill_volume(variant, volume_cell, stats):
    if variant.size and variant.size.volume_litres is None and volume_cell:
        try:
            variant.size.volume_litres = Decimal(str(volume_cell))
        except InvalidOperation:
            return
        variant.size.save(update_fields=['volume_litres'])
        stats['volume_backfills'] += 1


def _sync_dilutions(variant, cells, app_types):
    """Replace the variant's dilution set with the sheet row's values."""
    count = 0
    keep_ids = []
    for col, cell_value in cells.items():
        if cell_value is None:
            continue
        value, note = None, ''
        if isinstance(cell_value, (int, float)):
            value = Decimal(str(cell_value))
        else:
            note = str(cell_value).strip()
        dilution, _ = DilutionVariant.objects.update_or_create(
            variant=variant,
            application_type=app_types[col],
            defaults={'value': value, 'note': note},
        )
        keep_ids.append(dilution.pk)
        count += 1
    variant.dilutions.exclude(pk__in=keep_ids).delete()
    return count
